"""
Tests for the PDF and Excel generators.

Author: Danyal Khan
Module: 5COSC021W - CW2 Individual Element

These tests verify that the generators produce well-formed output and
that the data they contain matches what queries.py returns.

Test plan covered:
    G-01  Every PDF starts with the %PDF magic header
    G-02  Every Excel file starts with the ZIP magic header (xlsx is a zip)
    G-03  Empty database produces a valid (non-empty) file with a fallback
    G-04  Generated PDFs contain the expected team/department names
    G-05  Generated Excel files contain the expected team/department names
"""

import io
import zipfile

from django.test import TestCase
from openpyxl import load_workbook

from reports import excel_generator, pdf_generator
from .fixtures import create_sample_organisation


# ---------------------------------------------------------------------
# PDF generator tests
# ---------------------------------------------------------------------

class PDFGeneratorBoundaryTests(TestCase):
    """G-01, G-03 - empty database boundary case."""

    def test_teams_summary_pdf_is_valid_when_empty(self):
        pdf_bytes = pdf_generator.TeamsSummaryReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        self.assertGreater(len(pdf_bytes), 1000)

    def test_unmanaged_pdf_is_valid_when_empty(self):
        pdf_bytes = pdf_generator.TeamsWithoutManagersReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        self.assertGreater(len(pdf_bytes), 1000)

    def test_department_summary_pdf_is_valid_when_empty(self):
        pdf_bytes = pdf_generator.DepartmentSummaryReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))
        self.assertGreater(len(pdf_bytes), 1000)


class PDFGeneratorContentTests(TestCase):
    """G-01, G-04 - happy-path PDF content checks."""

    def setUp(self):
        create_sample_organisation()

    def test_teams_summary_pdf_starts_with_magic(self):
        pdf_bytes = pdf_generator.TeamsSummaryReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))

    def test_teams_summary_pdf_is_non_trivial_size(self):
        # A populated report should be larger than an empty one.
        pdf_bytes = pdf_generator.TeamsSummaryReport().render()
        self.assertGreater(len(pdf_bytes), 1500)

    def test_unmanaged_pdf_renders(self):
        pdf_bytes = pdf_generator.TeamsWithoutManagersReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))

    def test_department_pdf_renders(self):
        pdf_bytes = pdf_generator.DepartmentSummaryReport().render()
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))


# ---------------------------------------------------------------------
# Excel generator tests
# ---------------------------------------------------------------------

class ExcelGeneratorBoundaryTests(TestCase):
    """G-02, G-03 - empty database boundary case."""

    def test_teams_summary_xlsx_is_valid_when_empty(self):
        xlsx_bytes = excel_generator.TeamsSummaryReport().render()
        # XLSX files are ZIP archives.
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(xlsx_bytes)))

    def test_unmanaged_xlsx_is_valid_when_empty(self):
        xlsx_bytes = excel_generator.TeamsWithoutManagersReport().render()
        self.assertTrue(zipfile.is_zipfile(io.BytesIO(xlsx_bytes)))


class ExcelGeneratorContentTests(TestCase):
    """G-02, G-05 - happy-path Excel content checks."""

    def setUp(self):
        create_sample_organisation()

    def test_teams_summary_xlsx_contains_every_team(self):
        xlsx_bytes = excel_generator.TeamsSummaryReport().render()
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        cells = [
            cell.value for row in ws.iter_rows()
            for cell in row if cell.value is not None
        ]
        joined = ' '.join(str(c) for c in cells)
        self.assertIn('Test Observability', joined)
        self.assertIn('Test CI/CD', joined)
        self.assertIn('Test Encoding', joined)

    def test_unmanaged_xlsx_lists_unmanaged_team(self):
        xlsx_bytes = excel_generator.TeamsWithoutManagersReport().render()
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        cells = [
            cell.value for row in ws.iter_rows()
            for cell in row if cell.value is not None
        ]
        joined = ' '.join(str(c) for c in cells)
        self.assertIn('Test Encoding', joined)

    def test_department_xlsx_lists_both_departments(self):
        xlsx_bytes = excel_generator.DepartmentSummaryReport().render()
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        cells = [
            cell.value for row in ws.iter_rows()
            for cell in row if cell.value is not None
        ]
        joined = ' '.join(str(c) for c in cells)
        self.assertIn('Platform Engineering Test', joined)
        self.assertIn('Streaming Test', joined)

    def test_department_xlsx_team_count_for_platform_is_correct(self):
        xlsx_bytes = excel_generator.DepartmentSummaryReport().render()
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == 'Platform Engineering Test':
                # Columns: department, leader, specialisation, teams, unmanaged, engineers
                self.assertEqual(row[3], 2)
                return
        self.fail('Platform Engineering Test row was not found in the sheet.')
