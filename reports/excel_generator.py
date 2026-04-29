"""
Excel report generation using openpyxl.

Author: Danyal Khan
Module: 5COSC021W - CW2 Individual Element


"""

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from . import branding
from . import queries


def _hex_to_argb(hex_colour: str) -> str:
    """openpyxl requires ARGB ('FF' prefix) - convert from '#RRGGBB'."""
    return 'FF' + hex_colour.lstrip('#').upper()


class BaseExcelReport:
    """Common scaffolding for every Sky Engineering Portal Excel report."""

    title = 'Sky Engineering Report'
    sheet_name = 'Report'
    filename = 'sky_report.xlsx'

    # -- Public API -----------------------------------------------------

    def render(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        self._write_chrome(ws)
        self.write_body(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def write_body(self, ws):
        """Override in subclasses to write the report data."""
        raise NotImplementedError

    # -- Helpers exposed to subclasses ---------------------------------

    def write_table(self, ws, start_row: int,
                    header: Iterable[str], rows: list[list]) -> int:
        """Write a header row and data rows starting at start_row.

        Returns the row index of the next free row after the table.
        """
        header_list = list(header)

        # Header row styling
        header_fill = PatternFill(
            start_color=_hex_to_argb(branding.SKY_NAVY),
            end_color=_hex_to_argb(branding.SKY_NAVY),
            fill_type='solid',
        )
        header_font = Font(
            name='Calibri', size=11, bold=True,
            color=_hex_to_argb(branding.WHITE),
        )
        header_align = Alignment(
            horizontal='left', vertical='center', wrap_text=True,
        )

        for col_idx, value in enumerate(header_list, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Body rows
        body_font = Font(name='Calibri', size=10)
        body_align = Alignment(
            horizontal='left', vertical='center', wrap_text=True,
        )
        zebra_fill = PatternFill(
            start_color=_hex_to_argb(branding.SKY_LIGHT_BLUE),
            end_color=_hex_to_argb(branding.SKY_LIGHT_BLUE),
            fill_type='solid',
        )
        thin_border = Border(
            left=Side(style='thin', color=_hex_to_argb(branding.BORDER_GREY)),
            right=Side(style='thin', color=_hex_to_argb(branding.BORDER_GREY)),
            top=Side(style='thin', color=_hex_to_argb(branding.BORDER_GREY)),
            bottom=Side(style='thin', color=_hex_to_argb(branding.BORDER_GREY)),
        )

        for row_offset, row_data in enumerate(rows, start=1):
            current_row = start_row + row_offset
            apply_zebra = row_offset % 2 == 0

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border
                if apply_zebra:
                    cell.fill = zebra_fill

        # Auto-width columns based on content. Capped to keep the sheet
        # from getting silly-wide on long URLs.
        for col_idx, header_value in enumerate(header_list, start=1):
            column_letter = get_column_letter(col_idx)
            max_len = len(str(header_value))
            for row_data in rows:
                cell_value = str(row_data[col_idx - 1]) if col_idx - 1 < len(row_data) else ''
                if len(cell_value) > max_len:
                    max_len = len(cell_value)
            ws.column_dimensions[column_letter].width = min(max_len + 4, 50)

        # Freeze panes so the header stays visible while scrolling.
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        return start_row + len(rows) + 1

    # -- Internal helpers -----------------------------------------------

    def _write_chrome(self, ws):
        """Write the branded header rows above the data table."""
        # Row 1: Big title in Sky navy
        title_cell = ws.cell(row=1, column=1, value=f'sky  ·  {self.title}')
        title_cell.font = Font(
            name='Calibri', size=18, bold=True,
            color=_hex_to_argb(branding.SKY_NAVY),
        )
        ws.row_dimensions[1].height = 28

        # Row 2: Generation metadata in muted text
        totals = queries.overall_totals()
        meta_cell = ws.cell(
            row=2, column=1,
            value=f'Generated {totals["generated_at"]}  ·  '
                  f'{totals["total_teams"]} team(s) across '
                  f'{totals["total_departments"]} department(s)',
        )
        meta_cell.font = Font(
            name='Calibri', size=10, italic=True,
            color=_hex_to_argb(branding.TEXT_MUTED),
        )
        ws.row_dimensions[2].height = 18


# ---------------------------------------------------------------------
# Concrete reports
# ---------------------------------------------------------------------

class TeamsSummaryReport(BaseExcelReport):
    title = 'All Teams'
    sheet_name = 'Teams'
    filename = 'sky_teams_summary.xlsx'

    def write_body(self, ws):
        rows = queries.teams_summary()
        data = [
            [r['team'], r['department'], r['manager'],
             r['members'], r['contact'], r['repository']]
            for r in rows
        ]
        if not data:
            ws.cell(row=4, column=1,
                    value='No teams are currently recorded.')
            return

        self.write_table(
            ws,
            start_row=4,
            header=['Team', 'Department', 'Manager',
                    'Members', 'Contact', 'Repository'],
            rows=data,
        )


class TeamsWithoutManagersReport(BaseExcelReport):
    title = 'Teams Without Managers'
    sheet_name = 'Unmanaged Teams'
    filename = 'sky_teams_without_managers.xlsx'

    def write_body(self, ws):
        rows = queries.teams_without_managers()
        data = [
            [r['team'], r['department'], r['members'],
             r['contact'], r['created']]
            for r in rows
        ]
        if not data:
            ws.cell(row=4, column=1,
                    value='Every team currently has a manager assigned.')
            return

        self.write_table(
            ws,
            start_row=4,
            header=['Team', 'Department', 'Members',
                    'Contact', 'Created'],
            rows=data,
        )


class DepartmentSummaryReport(BaseExcelReport):
    title = 'Department Summary'
    sheet_name = 'Departments'
    filename = 'sky_department_summary.xlsx'

    def write_body(self, ws):
        rows = queries.department_summary()
        data = [
            [r['department'], r['leader'], r['specialisation'],
             r['num_teams'], r['unmanaged_teams'], r['unique_engineers']]
            for r in rows
        ]
        if not data:
            ws.cell(row=4, column=1,
                    value='No departments are currently recorded.')
            return

        self.write_table(
            ws,
            start_row=4,
            header=['Department', 'Leader', 'Specialisation',
                    'Teams', 'Unmanaged', 'Engineers'],
            rows=data,
        )
